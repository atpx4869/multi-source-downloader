# -*- coding: utf-8 -*-
"""
æ ‡å‡†æŸ¥æ–°å¯¹è¯æ¡†
ç”¨æˆ·ä¸Šä¼ æ ‡å‡†å·åˆ—è¡¨ï¼Œè‡ªåŠ¨æŸ¥è¯¢å‘å¸ƒæ—¥æœŸã€å®æ–½æ—¥æœŸã€çŠ¶æ€ç­‰å…ƒæ•°æ®ä¿¡æ¯
"""
import sys
import threading
from pathlib import Path
from datetime import datetime
import pandas as pd

project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from core.aggregated_downloader import AggregatedDownloader

try:
    from PySide6 import QtCore, QtWidgets, QtGui
    PYSIDE_VER = 6
except ImportError:
    from PySide2 import QtCore, QtWidgets, QtGui
    PYSIDE_VER = 2


class StandardInfoDialog(QtWidgets.QDialog):
    """æ ‡å‡†æŸ¥æ–°å¯¹è¯æ¡†"""
    
    def __init__(self, parent=None, parent_settings=None):
        super().__init__(parent)
        self.setWindowTitle("æ ‡å‡†æŸ¥æ–° - æ‰¹é‡æŸ¥è¯¢å…ƒæ•°æ®")
        self.setGeometry(50, 50, 1400, 800)  # åŠ å®½çª—å£ä»¥å®¹çº³æ›´å¤šå†…å®¹
        self.setModal(True)
        self.parent_settings = parent_settings  # ä¿å­˜ä¸»ç¨‹åºé…ç½®
        
        # è®¾ç½®å¯¹è¯æ¡†æ ·å¼
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #333333;
            }
            QGroupBox {
                color: #333333;
                font-weight: bold;
                border: 1px solid #ddd;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                color: #2196F3;
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 0 5px;
            }
            QTableWidget {
                background-color: white;
                color: #333333;
                gridline-color: #ddd;
                selection-background-color: #2196F3;
                selection-color: white;
            }
            QTableWidget::item {
                color: #333333;
            }
            QHeaderView::section {
                background-color: #e8e8e8;
                color: #333333;
                padding: 5px;
                border: 1px solid #ddd;
                font-weight: bold;
            }
        """)
        
        self.downloader = None
        self.result_df = None
        self.file_path = None
        self.processing = False
        
        self.init_ui()
    
    def init_ui(self):
        """åˆå§‹åŒ– UI"""
        main_layout = QtWidgets.QVBoxLayout()
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # ========== 1. æ–‡ä»¶é€‰æ‹©åŒº ==========
        file_group = QtWidgets.QGroupBox("1. é€‰æ‹©æ ‡å‡†å·æ–‡ä»¶")
        file_layout = QtWidgets.QHBoxLayout()
        
        self.label_file = QtWidgets.QLabel("æœªé€‰æ‹©æ–‡ä»¶")
        self.label_file.setStyleSheet("color: #666666; padding: 5px;")
        
        self.btn_select = QtWidgets.QPushButton("é€‰æ‹©æ–‡ä»¶ (Excel/CSV/TXT)")
        self.btn_select.setMaximumWidth(160)
        self.btn_select.clicked.connect(self.select_file)
        self.btn_select.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
        """)
        
        file_layout.addWidget(self.label_file, 1)
        file_layout.addWidget(self.btn_select)
        file_group.setLayout(file_layout)
        main_layout.addWidget(file_group)
        
        # ========== 2. é…ç½®åŒº ==========
        config_group = QtWidgets.QGroupBox("2. é…ç½®æŸ¥è¯¢å‚æ•°")
        config_layout = QtWidgets.QFormLayout()
        
        # æ•°æ®æºé€‰æ‹©
        source_layout = QtWidgets.QHBoxLayout()
        self.cb_zby = QtWidgets.QCheckBox("ZBY")
        self.cb_by = QtWidgets.QCheckBox("BY")
        self.cb_gbw = QtWidgets.QCheckBox("GBW")
        
        # åˆ·æ–°æŒ‰é’®
        self.btn_refresh_sources = QtWidgets.QPushButton("åˆ·æ–°")
        self.btn_refresh_sources.setMaximumWidth(60)
        self.btn_refresh_sources.clicked.connect(self.refresh_sources_from_parent)
        self.btn_refresh_sources.setStyleSheet("""
            QPushButton {
                background-color: #607D8B;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 5px 10px;
                font-size: 12px;
            }
            QPushButton:hover {
                background-color: #455A64;
            }
        """)
        
        # åˆå§‹åŒ–æ•°æ®æºé€‰æ‹©ï¼ˆä»ä¸»ç¨‹åºé…ç½®åŠ è½½ï¼‰
        self._init_sources_from_parent()
        
        source_layout.addWidget(self.cb_zby)
        source_layout.addWidget(self.cb_by)
        source_layout.addWidget(self.cb_gbw)
        source_layout.addWidget(self.btn_refresh_sources)
        source_layout.addStretch()
        config_layout.addRow("æ•°æ®æº:", source_layout)
        
        # åˆ—åç§°é…ç½®
        self.input_column = QtWidgets.QLineEdit()
        self.input_column.setPlaceholderText("ä¾‹å¦‚: æ ‡å‡†å·, std_no, æ ‡å‡†ç¼–å· (ç•™ç©ºåˆ™è‡ªåŠ¨è¯†åˆ«)")
        config_layout.addRow("æ ‡å‡†å·åˆ—å:", self.input_column)
        
        config_group.setLayout(config_layout)
        main_layout.addWidget(config_group)
        
        # ========== 3. å¤„ç†æŒ‰é’® ==========
        btn_layout = QtWidgets.QHBoxLayout()
        
        self.btn_process = QtWidgets.QPushButton("ğŸ” å¼€å§‹æŸ¥è¯¢")
        self.btn_process.setEnabled(False)
        self.btn_process.clicked.connect(self.process_file)
        self.btn_process.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 20px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_process)
        btn_layout.addStretch()
        main_layout.addLayout(btn_layout)
        
        # ========== 4. è¿›åº¦æ˜¾ç¤º ==========
        self.progress = QtWidgets.QProgressBar()
        self.progress.setVisible(False)
        self.progress.setStyleSheet("""
            QProgressBar {
                border: 1px solid #ddd;
                border-radius: 4px;
                text-align: center;
                background-color: white;
            }
            QProgressBar::chunk {
                background-color: #2196F3;
            }
        """)
        main_layout.addWidget(self.progress)
        
        self.label_status = QtWidgets.QLabel("")
        self.label_status.setStyleSheet("color: #666666; padding: 5px;")
        self.label_status.setAlignment(QtCore.Qt.AlignCenter)
        main_layout.addWidget(self.label_status)
        
        # ========== 5. ç»“æœæ˜¾ç¤ºè¡¨æ ¼ ==========
        result_group = QtWidgets.QGroupBox("3. æŸ¥è¯¢ç»“æœ")
        result_layout = QtWidgets.QVBoxLayout()
        
        self.table = QtWidgets.QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            'åŸå§‹æ ‡å‡†å·', 'è§„èŒƒæ ‡å‡†å·', 'æ ‡å‡†åç§°', 
            'å‘å¸ƒæ—¥æœŸ', 'å®æ–½æ—¥æœŸ', 'çŠ¶æ€', 
            'æ›¿ä»£æ ‡å‡†', 'æ›¿ä»£å®æ–½æ—¥æœŸ', 'æ›¿ä»£æ ‡å‡†åç§°',
            'æ˜¯å¦å˜æ›´'
        ])
        self.table.horizontalHeader().setStretchLastSection(False)
        # è®¾ç½®åˆ—å®½æ¨¡å¼ï¼šåŸå§‹æ ‡å‡†å·ã€è§„èŒƒæ ‡å‡†å·ã€æ›¿ä»£æ ‡å‡†è‡ªé€‚åº”å†…å®¹ï¼Œæ ‡å‡†åç§°ã€æ›¿ä»£åç§°æ‹‰ä¼¸
        self.table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(6, QtWidgets.QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(8, QtWidgets.QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        
        result_layout.addWidget(self.table)
        result_group.setLayout(result_layout)
        main_layout.addWidget(result_group, 1)
        
        # ========== 6. åº•éƒ¨æŒ‰é’® ==========
        bottom_layout = QtWidgets.QHBoxLayout()
        
        self.btn_export_excel = QtWidgets.QPushButton("å¯¼å‡º Excel")
        self.btn_export_excel.setMaximumWidth(100)
        self.btn_export_excel.setEnabled(False)
        self.btn_export_excel.clicked.connect(self.export_excel)
        self.btn_export_excel.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        
        self.btn_export_csv = QtWidgets.QPushButton("å¯¼å‡º CSV")
        self.btn_export_csv.setMaximumWidth(100)
        self.btn_export_csv.setEnabled(False)
        self.btn_export_csv.clicked.connect(self.export_csv)
        self.btn_export_csv.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        
        self.btn_close = QtWidgets.QPushButton("å…³é—­")
        self.btn_close.setMaximumWidth(80)
        self.btn_close.clicked.connect(self.close)
        self.btn_close.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 15px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        
        bottom_layout.addStretch()
        bottom_layout.addWidget(self.btn_export_excel)
        bottom_layout.addWidget(self.btn_export_csv)
        bottom_layout.addWidget(self.btn_close)
        
        main_layout.addLayout(bottom_layout)
        self.setLayout(main_layout)
    
    def select_file(self):
        """é€‰æ‹©æ–‡ä»¶"""
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "é€‰æ‹©æ ‡å‡†å·æ–‡ä»¶",
            "",
            "æ‰€æœ‰æ”¯æŒçš„æ–‡ä»¶ (*.xlsx *.xls *.csv *.txt);;Excelæ–‡ä»¶ (*.xlsx *.xls);;CSVæ–‡ä»¶ (*.csv);;æ–‡æœ¬æ–‡ä»¶ (*.txt)"
        )
        
        if file_path:
            self.file_path = file_path
            self.label_file.setText(f"å·²é€‰æ‹©: {Path(file_path).name}")
            self.label_file.setStyleSheet("color: #4CAF50; padding: 5px; font-weight: bold;")
            self.btn_process.setEnabled(True)
    
    def _init_sources_from_parent(self):
        """ä»ä¸»ç¨‹åºé…ç½®åˆå§‹åŒ–æ•°æ®æºé€‰æ‹©"""
        if self.parent_settings and 'sources' in self.parent_settings:
            parent_sources = self.parent_settings['sources']
            self.cb_zby.setChecked('ZBY' in parent_sources)
            self.cb_by.setChecked('BY' in parent_sources)
            self.cb_gbw.setChecked('GBW' in parent_sources)
        else:
            # é»˜è®¤åªé€‰æ‹©ZBY
            self.cb_zby.setChecked(True)
            self.cb_by.setChecked(False)
            self.cb_gbw.setChecked(False)
    
    def refresh_sources_from_parent(self):
        """åˆ·æ–°æ•°æ®æºé€‰æ‹©ï¼ˆä»ä¸»ç¨‹åºé…ç½®ï¼‰"""
        self._init_sources_from_parent()
        QtWidgets.QMessageBox.information(self, "æç¤º", "å·²åŒæ­¥ä¸»ç¨‹åºçš„æ•°æ®æºé…ç½®")
    
    def get_enabled_sources(self):
        """è·å–å¯ç”¨çš„æ•°æ®æº"""
        sources = []
        if self.cb_zby.isChecked():
            sources.append("ZBY")
        if self.cb_by.isChecked():
            sources.append("BY")
        if self.cb_gbw.isChecked():
            sources.append("GBW")
        return sources if sources else ["ZBY"]
    
    def process_file(self):
        """å¤„ç†æ–‡ä»¶"""
        if self.processing:
            QtWidgets.QMessageBox.warning(self, "æç¤º", "æ­£åœ¨å¤„ç†ä¸­ï¼Œè¯·ç¨å€™...")
            return
        
        if not self.file_path:
            QtWidgets.QMessageBox.warning(self, "æç¤º", "è¯·å…ˆé€‰æ‹©æ–‡ä»¶ï¼")
            return
        
        self.processing = True
        self.btn_process.setEnabled(False)
        self.progress.setVisible(True)
        self.progress.setValue(0)
        self.label_status.setText("æ­£åœ¨è¯»å–æ–‡ä»¶...")
        
        # åœ¨åå°çº¿ç¨‹å¤„ç†
        thread = threading.Thread(target=self._process_in_background)
        thread.daemon = True
        thread.start()
    
    def _process_in_background(self):
        """åå°å¤„ç†çº¿ç¨‹"""
        try:
            # 1. è¯»å–æ–‡ä»¶
            self.update_status("æ­£åœ¨è¯»å–æ–‡ä»¶...", 10)
            df = self._read_file()
            
            if df is None or df.empty:
                self.update_status("æ–‡ä»¶è¯»å–å¤±è´¥æˆ–ä¸ºç©ºï¼", 0)
                self.processing = False
                return
            
            # 2. è¯†åˆ«æ ‡å‡†å·åˆ—
            self.update_status("æ­£åœ¨è¯†åˆ«æ ‡å‡†å·åˆ—...", 20)
            std_column = self._identify_std_column(df)
            
            if not std_column:
                self.update_status("æœªæ‰¾åˆ°æ ‡å‡†å·åˆ—ï¼", 0)
                self.processing = False
                return
            
            # 3. æå–æ ‡å‡†å·
            std_list = df[std_column].dropna().astype(str).tolist()
            self.update_status(f"æ‰¾åˆ° {len(std_list)} ä¸ªæ ‡å‡†å·", 30)
            
            # 4. åˆå§‹åŒ–ä¸‹è½½å™¨
            sources = self.get_enabled_sources()
            self.update_status(f"åˆå§‹åŒ–æ•°æ®æº: {', '.join(sources)}", 40)
            self.downloader = AggregatedDownloader(enable_sources=sources, output_dir="downloads")
            
            # 5. æ‰¹é‡æŸ¥è¯¢
            results = []
            total = len(std_list)
            for i, std_no in enumerate(std_list, 1):
                progress = 40 + int((i / total) * 50)
                self.update_status(f"æŸ¥è¯¢ä¸­ ({i}/{total}): {std_no}", progress)
                
                # æœç´¢æ ‡å‡† - å¢åŠ limitä»¥ä¾¿æ™ºèƒ½ç­›é€‰
                import re
                has_year = bool(re.search(r'-\d{4}$', std_no.strip()))
                
                # æå–å¹´ä»½çš„è¾…åŠ©å‡½æ•°
                def extract_year(r):
                    match = re.search(r'-(\d{4})$', r.std_no)
                    return int(match.group(1)) if match else 0
                
                # å…ˆæœç´¢ç²¾ç¡®åŒ¹é…
                search_results = self.downloader.search(std_no, limit=10)
                
                # å¦‚æœå¸¦å¹´ä»£å·ï¼Œé¢å¤–æœç´¢ä¸å¸¦å¹´ä»£å·çš„ç‰ˆæœ¬ä»¥æŸ¥æ‰¾æ›´æ–°ç‰ˆæœ¬
                all_versions = []
                if has_year and search_results:
                    base_std_no = re.sub(r'-\d{4}$', '', std_no.strip())
                    all_versions = self.downloader.search(base_std_no, limit=10)
                
                if search_results:
                    # æ™ºèƒ½é€‰æ‹©ï¼šå¦‚æœä¸å¸¦å¹´ä»£å·ï¼Œä¼˜å…ˆé€‰æ‹©ç°è¡Œæ ‡å‡†
                    
                    # æå–å¹´ä»½çš„è¾…åŠ©å‡½æ•°
                    def extract_year(r):
                        match = re.search(r'-(\d{4})$', r.std_no)
                        return int(match.group(1)) if match else 0
                    
                    if not has_year and len(search_results) > 1:
                        # ä¸å¸¦å¹´ä»£å·ï¼Œä¼˜å…ˆé€‰æ‹©ç°è¡Œæ ‡å‡†
                        current_results = [r for r in search_results if r.status and "ç°è¡Œ" in r.status]
                        if current_results:
                            # å¦‚æœæœ‰å¤šä¸ªç°è¡Œæ ‡å‡†ï¼Œé€‰æ‹©å¹´ä»½æœ€æ–°çš„
                            current_results.sort(key=extract_year, reverse=True)
                            result = current_results[0]
                        else:
                            # æ²¡æœ‰ç°è¡Œæ ‡å‡†ï¼Œé€‰æ‹©å¹´ä»½æœ€æ–°çš„
                            search_results.sort(key=extract_year, reverse=True)
                            result = search_results[0]
                    else:
                        # å¸¦å¹´ä»£å·æˆ–åªæœ‰ä¸€ä¸ªç»“æœï¼Œç›´æ¥ä½¿ç”¨
                        result = search_results[0]
                    
                    # æ™ºèƒ½åˆ¤æ–­æ›¿ä»£æ ‡å‡†å’Œå˜æ›´çŠ¶æ€
                    replace_std_text = result.replace_std or ''
                    replace_implement = ''
                    replace_name = ''
                    is_changed = ''
                    
                    # æƒ…å†µ1ï¼šAPIæ˜ç¡®è¿”å›äº†æ›¿ä»£æ ‡å‡†
                    if replace_std_text.strip():
                        is_changed = 'å˜æ›´'
                    
                    # æƒ…å†µ2å’Œ3ï¼šå°è¯•é€šè¿‡æœç´¢æ‰€æœ‰ç‰ˆæœ¬æ¥æŸ¥æ‰¾æ›¿ä»£æ ‡å‡†
                    # æ— è®ºæ˜¯å¦å¸¦å¹´ä»£å·ï¼Œéƒ½å°è¯•æŸ¥æ‰¾æ›´æ–°ç‰ˆæœ¬
                    if not replace_std_text.strip():  # åªæœ‰åœ¨APIæ²¡æœ‰è¿”å›æ›¿ä»£æ ‡å‡†æ—¶æ‰æŸ¥æ‰¾
                        # è·å–æ‰€æœ‰ç‰ˆæœ¬ï¼ˆå¦‚æœè¿˜æ²¡è·å–ï¼‰
                        if not all_versions:
                            base_std_no = re.sub(r'-\d{4}$', '', result.std_no.strip())
                            all_versions = self.downloader.search(base_std_no, limit=10)
                        
                        if all_versions and len(all_versions) > 1:
                            # åªçœ‹åŸºç¡€æ ‡å‡†å·å®Œå…¨ç›¸åŒçš„ç‰ˆæœ¬
                            current_year = extract_year(result)
                            base_std_no = re.sub(r'-\d{4}$', '', result.std_no.strip())
                            
                            # è¿‡æ»¤å‡ºåŸºç¡€æ ‡å‡†å·å®Œå…¨ç›¸åŒçš„ç‰ˆæœ¬
                            same_base_versions = [
                                r for r in all_versions 
                                if re.sub(r'-\d{4}$', '', r.std_no.strip()) == base_std_no
                            ]
                            
                            # åœ¨åŒä¸€æ ‡å‡†å·çš„ç‰ˆæœ¬ä¸­æŸ¥æ‰¾æ›´æ–°çš„
                            newer_versions = [r for r in same_base_versions if extract_year(r) > current_year]
                            if newer_versions:
                                # æ‰¾åˆ°äº†æ›´æ–°çš„ç‰ˆæœ¬
                                newer_versions.sort(key=extract_year, reverse=True)
                                newest = newer_versions[0]
                                replace_std_text = newest.std_no
                                replace_implement = newest.implement or ''
                                replace_name = newest.name or ''
                                is_changed = 'å˜æ›´'
                    
                    # å¦‚æœçŠ¶æ€ä¸ºåºŸæ­¢ï¼Œç¡®ä¿æ ‡è®°ä¸ºå˜æ›´ï¼ˆå³ä½¿æ²¡æ‰¾åˆ°æ›¿ä»£æ ‡å‡†ï¼‰
                    if result.status and ('åºŸæ­¢' in result.status or 'å³å°†åºŸæ­¢' in result.status):
                        is_changed = 'å˜æ›´'
                    
                    results.append({
                        'åŸå§‹æ ‡å‡†å·': std_no,
                        'è§„èŒƒæ ‡å‡†å·': result.std_no,
                        'æ ‡å‡†åç§°': result.name,
                        'å‘å¸ƒæ—¥æœŸ': result.publish or '',
                        'å®æ–½æ—¥æœŸ': result.implement or '',
                        'çŠ¶æ€': result.status or '',
                        'æ›¿ä»£æ ‡å‡†': replace_std_text,
                        'æ›¿ä»£å®æ–½æ—¥æœŸ': replace_implement,
                        'æ›¿ä»£æ ‡å‡†åç§°': replace_name,
                        'æ˜¯å¦å˜æ›´': is_changed
                    })
                else:
                    results.append({
                        'åŸå§‹æ ‡å‡†å·': std_no,
                        'è§„èŒƒæ ‡å‡†å·': '',
                        'æ ‡å‡†åç§°': 'æœªæ‰¾åˆ°',
                        'å‘å¸ƒæ—¥æœŸ': '',
                        'å®æ–½æ—¥æœŸ': '',
                        'çŠ¶æ€': '',
                        'æ›¿ä»£æ ‡å‡†': '',
                        'æ›¿ä»£å®æ–½æ—¥æœŸ': '',
                        'æ›¿ä»£æ ‡å‡†åç§°': '',
                        'æ˜¯å¦å˜æ›´': ''
                    })
            
            # 6. æ˜¾ç¤ºç»“æœ
            self.result_df = pd.DataFrame(results)
            self.update_status(f"æŸ¥è¯¢å®Œæˆï¼æ‰¾åˆ° {len(results)} æ¡ç»“æœ", 100)
            self._display_results()
            
        except Exception as e:
            self.update_status(f"å¤„ç†å¤±è´¥: {str(e)}", 0)
            QtWidgets.QMessageBox.critical(self, "é”™è¯¯", f"å¤„ç†å¤±è´¥:\n{str(e)}")
        
        finally:
            self.processing = False
            self.btn_process.setEnabled(True)
    
    def _read_file(self):
        """è¯»å–æ–‡ä»¶"""
        try:
            file_ext = Path(self.file_path).suffix.lower()
            
            if file_ext in ['.xlsx', '.xls']:
                return pd.read_excel(self.file_path)
            elif file_ext == '.csv':
                # å°è¯•ä¸åŒç¼–ç 
                for encoding in ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']:
                    try:
                        return pd.read_csv(self.file_path, encoding=encoding)
                    except:
                        continue
                return None
            elif file_ext == '.txt':
                # è¯»å–ä¸ºå•åˆ—
                with open(self.file_path, 'r', encoding='utf-8') as f:
                    lines = [line.strip() for line in f if line.strip()]
                return pd.DataFrame({'æ ‡å‡†å·': lines})
            else:
                return None
        except Exception as e:
            print(f"è¯»å–æ–‡ä»¶å¤±è´¥: {e}")
            return None
    
    def _identify_std_column(self, df):
        """è¯†åˆ«æ ‡å‡†å·åˆ—"""
        # ç”¨æˆ·æŒ‡å®šçš„åˆ—å
        user_column = self.input_column.text().strip()
        if user_column and user_column in df.columns:
            return user_column
        
        # è‡ªåŠ¨è¯†åˆ«
        possible_names = ['æ ‡å‡†å·', 'std_no', 'æ ‡å‡†ç¼–å·', 'ç¼–å·', 'number', 'code', 'æ ‡å‡†ä»£å·']
        for col in df.columns:
            if any(name in str(col).lower() for name in possible_names):
                return col
        
        # é»˜è®¤ä½¿ç”¨ç¬¬ä¸€åˆ—
        return df.columns[0] if len(df.columns) > 0 else None
    
    def _display_results(self):
        """æ˜¾ç¤ºç»“æœ"""
        if self.result_df is None or self.result_df.empty:
            return
        
        self.table.setRowCount(0)
        for idx, row in self.result_df.iterrows():
            row_pos = self.table.rowCount()
            self.table.insertRow(row_pos)
            
            for col_idx, col_name in enumerate(self.result_df.columns):
                value = str(row[col_name])
                item = QtWidgets.QTableWidgetItem(value)
                
                # çŠ¶æ€åˆ—é¢œè‰²
                if col_name == 'çŠ¶æ€' and value:
                    if 'ç°è¡Œ' in value or 'active' in value.lower():
                        item.setBackground(QtGui.QColor("#d4edda"))
                        item.setForeground(QtGui.QColor("#155724"))
                    elif 'åºŸæ­¢' in value or 'supersede' in value.lower():
                        item.setBackground(QtGui.QColor("#f8d7da"))
                        item.setForeground(QtGui.QColor("#721c24"))
                
                # æ˜¯å¦å˜æ›´åˆ—é¢œè‰²ï¼ˆæµ…è“è‰²ï¼‰
                if col_name == 'æ˜¯å¦å˜æ›´' and value == 'å˜æ›´':
                    item.setBackground(QtGui.QColor("#cfe2ff"))  # æµ…è“è‰²
                    item.setForeground(QtGui.QColor("#084298"))  # æ·±è“è‰²æ–‡å­—
                
                self.table.setItem(row_pos, col_idx, item)
        
        self.btn_export_excel.setEnabled(True)
        self.btn_export_csv.setEnabled(True)
    
    def update_status(self, message, progress):
        """æ›´æ–°çŠ¶æ€ï¼ˆçº¿ç¨‹å®‰å…¨ï¼‰"""
        QtCore.QMetaObject.invokeMethod(
            self.label_status,
            "setText",
            QtCore.Qt.QueuedConnection,
            QtCore.Q_ARG(str, message)
        )
        QtCore.QMetaObject.invokeMethod(
            self.progress,
            "setValue",
            QtCore.Qt.QueuedConnection,
            QtCore.Q_ARG(int, progress)
        )
    
    def export_excel(self):
        """å¯¼å‡ºä¸º Excelï¼ˆåŒ…å«æ ¼å¼ï¼‰"""
        if self.result_df is None or self.result_df.empty:
            QtWidgets.QMessageBox.warning(self, "æç¤º", "æš‚æ— ç»“æœå¯å¯¼å‡ºï¼")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"æ ‡å‡†æŸ¥æ–°ç»“æœ_{timestamp}.xlsx"
        
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "å¯¼å‡º Excel", default_name, "Excel æ–‡ä»¶ (*.xlsx)"
        )
        
        if file_path:
            try:
                # å¯¼å‡ºåˆ°Excel
                self.result_df.to_excel(file_path, index=False, engine='openpyxl')
                
                # ä½¿ç”¨openpyxlæ·»åŠ æ ¼å¼
                from openpyxl import load_workbook
                from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
                
                workbook = load_workbook(file_path)
                worksheet = workbook.active
                
                # å®šä¹‰è¾¹æ¡†æ ·å¼
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # å®šä¹‰å¡«å……è‰²ï¼ˆæµ…è“è‰²ï¼Œä¸UIä¸€è‡´ï¼‰
                light_blue_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
                
                # å®šä¹‰çŠ¶æ€é¢œè‰²ï¼ˆä¸UIä¿æŒä¸€è‡´ï¼‰
                # ç°è¡Œ - æµ…ç»¿è‰²
                status_active_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                status_active_font = Font(color="155724", bold=False)
                # åºŸæ­¢ - æµ…çº¢è‰²
                status_obsolete_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                status_obsolete_font = Font(color="721C24", bold=False)
                
                # éå†æ‰€æœ‰å•å…ƒæ ¼ï¼Œæ·»åŠ è¾¹æ¡†ã€è®¾ç½®è¡Œé«˜å’Œå¯¹é½æ–¹å¼
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # æŸ¥æ‰¾"æ˜¯å¦å˜æ›´"å’Œ"çŠ¶æ€"åˆ—çš„ä½ç½®
                change_col_idx = None
                status_col_idx = None
                if max_row > 0:
                    for col in range(1, max_col + 1):
                        header_cell = worksheet.cell(row=1, column=col)
                        if header_cell.value == 'æ˜¯å¦å˜æ›´':
                            change_col_idx = col
                        elif header_cell.value == 'çŠ¶æ€':
                            status_col_idx = col

                
                for row in range(1, max_row + 1):
                    # è®¾ç½®è¡Œé«˜ä¸º14
                    worksheet.row_dimensions[row].height = 14
                    
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        # æ·»åŠ è¾¹æ¡†
                        cell.border = thin_border
                        # è®¾ç½®å¯¹é½æ–¹å¼å’Œè‡ªåŠ¨æ¢è¡Œ
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                        
                        # è·³è¿‡è¡¨å¤´è¡Œ
                        if row == 1:
                            continue
                        
                        # æ ¹æ®çŠ¶æ€åˆ—çš„å€¼åº”ç”¨ä¸åŒçš„å¡«å……è‰²
                        if col == status_col_idx and cell.value:
                            cell_value = str(cell.value)
                            if 'ç°è¡Œ' in cell_value or 'active' in cell_value.lower():
                                cell.fill = status_active_fill
                                cell.font = status_active_font
                            elif 'åºŸæ­¢' in cell_value or 'supersede' in cell_value.lower():
                                cell.fill = status_obsolete_fill
                                cell.font = status_obsolete_font
                        
                        # å¦‚æœæ˜¯"æ˜¯å¦å˜æ›´"åˆ—ä¸”å€¼ä¸º"å˜æ›´"ï¼Œæ·»åŠ å¡«å……è‰²
                        if col == change_col_idx and cell.value == 'å˜æ›´':
                            cell.fill = light_blue_fill
                
                # è‡ªé€‚åº”åˆ—å®½
                for col_idx in range(1, max_col + 1):
                    max_length = 0
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    
                    for row in range(1, max_row + 1):
                        cell = worksheet.cell(row=row, column=col_idx)
                        try:
                            if cell.value:
                                # è®¡ç®—æ–‡æœ¬é•¿åº¦ï¼ˆä¸­æ–‡æŒ‰2ä¸ªå­—ç¬¦è®¡ç®—ï¼‰
                                cell_length = len(str(cell.value))
                                for char in str(cell.value):
                                    if ord(char) > 127:  # ä¸­æ–‡å­—ç¬¦
                                        cell_length += 1
                                max_length = max(max_length, cell_length)
                        except:
                            pass
                    
                    # è®¾ç½®åˆ—å®½ï¼ˆåŠ ä¸Šä¸€ç‚¹ä½™é‡ï¼‰
                    adjusted_width = min(max_length + 2, 50)  # æœ€å¤§å®½åº¦50
                    worksheet.column_dimensions[col_letter].width = adjusted_width
                
                workbook.save(file_path)
                QtWidgets.QMessageBox.information(
                    self, "æˆåŠŸ", f"å·²æˆåŠŸå¯¼å‡ºåˆ°:\n{file_path}"
                )
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "é”™è¯¯", f"å¯¼å‡ºå¤±è´¥:\n{str(e)}")
    
    def export_csv(self):
        """å¯¼å‡ºä¸º CSV"""
        if self.result_df is None or self.result_df.empty:
            QtWidgets.QMessageBox.warning(self, "æç¤º", "æš‚æ— ç»“æœå¯å¯¼å‡ºï¼")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"æ ‡å‡†æŸ¥æ–°ç»“æœ_{timestamp}.csv"
        
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "å¯¼å‡º CSV", default_name, "CSV æ–‡ä»¶ (*.csv)"
        )
        
        if file_path:
            try:
                self.result_df.to_csv(file_path, index=False, encoding='utf-8-sig')
                QtWidgets.QMessageBox.information(
                    self, "æˆåŠŸ", f"å·²æˆåŠŸå¯¼å‡ºåˆ°:\n{file_path}"
                )
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, "é”™è¯¯", f"å¯¼å‡ºå¤±è´¥:\n{str(e)}")


if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    dialog = StandardInfoDialog()
    dialog.show()
    sys.exit(app.exec())
